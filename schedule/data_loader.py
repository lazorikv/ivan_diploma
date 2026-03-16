from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent

UPLOADS_DIR = BASE_DIR / "uploads"

_DEFAULT_GROUPS_XLSX = BASE_DIR / "Группы.xlsx"
_DEFAULT_TEACHERS_XLSX = BASE_DIR / "Преподаватели и дисциплины.xlsx"

_UPLOADED_GROUPS_XLSX = UPLOADS_DIR / "groups.xlsx"
_UPLOADED_TEACHERS_XLSX = UPLOADS_DIR / "teachers.xlsx"


def _resolve_groups_path() -> Path:
    return _UPLOADED_GROUPS_XLSX if _UPLOADED_GROUPS_XLSX.exists() else _DEFAULT_GROUPS_XLSX


def _resolve_teachers_path() -> Path:
    return _UPLOADED_TEACHERS_XLSX if _UPLOADED_TEACHERS_XLSX.exists() else _DEFAULT_TEACHERS_XLSX


def get_upload_status() -> dict:
    """Return info about which xlsx files are currently in use."""
    groups_uploaded = _UPLOADED_GROUPS_XLSX.exists()
    teachers_uploaded = _UPLOADED_TEACHERS_XLSX.exists()
    return {
        "groups_uploaded": groups_uploaded,
        "groups_path": str(_UPLOADED_GROUPS_XLSX if groups_uploaded else _DEFAULT_GROUPS_XLSX),
        "groups_filename": _UPLOADED_GROUPS_XLSX.name if groups_uploaded else _DEFAULT_GROUPS_XLSX.name,
        "teachers_uploaded": teachers_uploaded,
        "teachers_path": str(_UPLOADED_TEACHERS_XLSX if teachers_uploaded else _DEFAULT_TEACHERS_XLSX),
        "teachers_filename": _UPLOADED_TEACHERS_XLSX.name if teachers_uploaded else _DEFAULT_TEACHERS_XLSX.name,
    }


# Keep module-level names for backward compatibility
GROUPS_XLSX = _DEFAULT_GROUPS_XLSX
TEACHERS_XLSX = _DEFAULT_TEACHERS_XLSX

PFK_DISC = "ПФК"
PFK_FULL = "Прикладная физическая культура"
PFK_TEACHER = "ПФК"
GROUP_SPLIT_THRESHOLD = 18


def _year_prefix(group_code: str) -> str:
    if group_code.startswith("М"):
        return "М1"
    return group_code[0] if group_code else "?"


@dataclass
class Lesson:
    group: str               # real group code OR "STREAM_..." for merged lessons
    discipline_short: str
    discipline_full: str
    lesson_type: str         # 'lec' | 'prac' | 'lab'
    teacher: str
    subgroup: int = 0
    count: int = 1
    # Non-empty: merged stream lesson — all groups attend the same slot
    stream_groups: list[str] = field(default_factory=list)


@dataclass
class GroupInfo:
    code: str
    size: Optional[int] = None


@dataclass
class ScheduleData:
    groups: list[GroupInfo]
    lessons: list[Lesson]
    disciplines: dict[str, str]


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_data() -> ScheduleData:
    groups_wb = openpyxl.load_workbook(_resolve_groups_path(), data_only=True)
    teachers_wb = openpyxl.load_workbook(_resolve_teachers_path(), data_only=True)

    # Sheet 1: group codes and sizes
    groups_ws = groups_wb["Лист1"]
    group_sizes: dict[str, Optional[int]] = {}
    for row in groups_ws.iter_rows(min_row=2, values_only=True):
        code = _clean(row[0])
        size = row[1] if len(row) > 1 else None
        if code:
            group_sizes[code] = int(size) if size else None

    # Sheet 2: hours per discipline per group
    hours_ws = groups_wb["Лист2"]
    hours_map: dict[tuple[str, str], dict[str, int]] = {}
    for row in hours_ws.iter_rows(min_row=2, values_only=True):
        grp = _clean(row[0])
        disc = _clean(row[1])
        lec = int(row[2]) if row[2] else 0
        prac = int(row[3]) if row[3] else 0
        lab = int(row[4]) if row[4] else 0
        if grp and disc:
            hours_map[(grp, disc)] = {"lec": lec, "prac": prac, "lab": lab}

    # Sheet 1 of teachers xlsx
    teachers_ws = teachers_wb["Лист1"]
    teacher_map: dict[tuple[str, str], dict] = {}
    disciplines: dict[str, str] = {}
    for row in teachers_ws.iter_rows(min_row=2, values_only=True):
        grp = _clean(row[0])
        disc_short = _clean(row[1])
        disc_full = _clean(row[2])
        t_lec = _clean(row[3])
        t_prac = _clean(row[4])
        t_lab1 = _clean(row[5]) if len(row) > 5 else ""
        t_lab2 = _clean(row[6]) if len(row) > 6 else ""
        if grp and disc_short:
            teacher_map[(grp, disc_short)] = {
                "lec": t_lec,
                "prac": t_prac,
                "lab1": t_lab1,
                "lab2": t_lab2,
            }
            if disc_short and disc_full:
                disciplines[disc_short] = disc_full

    # -----------------------------------------------------------
    # Merging rules:
    #   - Lectures: merge groups with same (teacher, disc, hours)
    #   - ПФК (practice, no teacher): merge ALL groups of the SAME YEAR into
    #     one stream — entire course year attends ПФК on the same day/slot
    #   - All other practices and labs: individual per group
    #   - Lab subgroups (split): always individual, never merged
    # -----------------------------------------------------------

    # (teacher, disc, 'lec', hours) → list of groups
    lec_merge: dict[tuple, list[str]] = defaultdict(list)
    lec_disc_full: dict[tuple, str] = {}

    # ПФК: year_prefix → list of groups in that year
    pfk_by_year: dict[str, list[str]] = defaultdict(list)
    pfk_hours_by_year: dict[str, int] = {}

    individual_lessons: list[Lesson] = []

    for (grp, disc), hours in hours_map.items():
        t_info = teacher_map.get((grp, disc), {})
        size = group_sizes.get(grp)
        needs_split = size is not None and size >= GROUP_SPLIT_THRESHOLD
        disc_full = disciplines.get(disc, disc)

        lec_h = hours.get("lec", 0)
        prac_h = hours.get("prac", 0)
        lab_h = hours.get("lab", 0)

        # --- Lectures: always try to merge ---
        if lec_h > 0:
            t = t_info.get("lec", "")
            if t and t != ".":
                key = (t, disc, "lec", lec_h)
                lec_merge[key].append(grp)
                lec_disc_full[key] = disc_full

        # --- ПФК practice: merge by YEAR (1 курс = один поток) ---
        if disc == PFK_DISC and prac_h > 0:
            year = _year_prefix(grp)
            pfk_by_year[year].append(grp)
            pfk_hours_by_year[year] = prac_h

        # --- Regular practices: individual per group ---
        if disc != PFK_DISC and prac_h > 0:
            t = t_info.get("prac", "")
            if t and t != ".":
                individual_lessons.append(Lesson(
                    group=grp,
                    discipline_short=disc,
                    discipline_full=disc_full,
                    lesson_type="prac",
                    teacher=t,
                    subgroup=0,
                    count=prac_h,
                ))

        # --- Labs: split → subgroups; non-split → individual ---
        if lab_h > 0:
            t1 = t_info.get("lab1", "")
            t2 = t_info.get("lab2", "")
            if needs_split and t1 and t1 != ".":
                t2_eff = t2 if (t2 and t2 != ".") else t1
                individual_lessons.append(Lesson(
                    group=grp,
                    discipline_short=disc,
                    discipline_full=disc_full,
                    lesson_type="lab",
                    teacher=t1,
                    subgroup=1,
                    count=lab_h,
                ))
                individual_lessons.append(Lesson(
                    group=grp,
                    discipline_short=disc,
                    discipline_full=disc_full,
                    lesson_type="lab",
                    teacher=t2_eff,
                    subgroup=2,
                    count=lab_h,
                ))
            elif t1 and t1 != ".":
                individual_lessons.append(Lesson(
                    group=grp,
                    discipline_short=disc,
                    discipline_full=disc_full,
                    lesson_type="lab",
                    teacher=t1,
                    subgroup=0,
                    count=lab_h,
                ))

    lessons: list[Lesson] = []

    # Build lecture lessons (merged if multiple groups)
    for (teacher, disc, ltype, count), grp_list in lec_merge.items():
        disc_full = lec_disc_full.get((teacher, disc, ltype, count), disc)
        unique = sorted(set(grp_list))
        if len(unique) > 1:
            lessons.append(Lesson(
                group=f"STREAM_{disc}_{teacher[:6]}",
                discipline_short=disc,
                discipline_full=disc_full,
                lesson_type="lec",
                teacher=teacher,
                subgroup=0,
                count=count,
                stream_groups=unique,
            ))
        else:
            lessons.append(Lesson(
                group=unique[0],
                discipline_short=disc,
                discipline_full=disc_full,
                lesson_type="lec",
                teacher=teacher,
                subgroup=0,
                count=count,
            ))

    # Build ПФК stream lessons — one stream per year.
    # count = hours from plan (should be 2: once per week → 2 sessions over 2 weeks).
    for year, grp_list in pfk_by_year.items():
        unique = sorted(set(grp_list))
        pfk_count = pfk_hours_by_year.get(year, 2)
        stream_id = f"STREAM_ПФК_{year}"
        lessons.append(Lesson(
            group=stream_id,
            discipline_short=PFK_DISC,
            discipline_full=PFK_FULL,
            lesson_type="prac",
            teacher=PFK_TEACHER,
            subgroup=0,
            count=pfk_count,
            stream_groups=unique,
        ))

    lessons.extend(individual_lessons)

    groups = [GroupInfo(code=c, size=s) for c, s in group_sizes.items()]

    return ScheduleData(groups=groups, lessons=lessons, disciplines=disciplines)


def expand_lessons(lessons: list[Lesson]) -> list[Lesson]:
    """Expand each lesson by its count into individual single-session entries."""
    expanded: list[Lesson] = []
    for lesson in lessons:
        for _ in range(lesson.count):
            expanded.append(Lesson(
                group=lesson.group,
                discipline_short=lesson.discipline_short,
                discipline_full=lesson.discipline_full,
                lesson_type=lesson.lesson_type,
                teacher=lesson.teacher,
                subgroup=lesson.subgroup,
                count=1,
                stream_groups=lesson.stream_groups[:],
            ))
    return expanded
